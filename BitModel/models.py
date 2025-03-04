'''
This Code main function is to calculte the ROP from the 4 following models:
1- Detournay et al. model
2- Ziaja's model
3- Gerbaud et al. model
4- Che et al. mode

It is also capable of calculating the WOB & TOB using Detournay et al. and Ziaja's models.

It can be utilized easily for studying and comparison purposes.

## Gerbaud et al. & Che et al. models are both single cutter models and integrated into full bit model using the number of cutters (nc)
'''

'''
Parameters used throughout the code:
omega (rev/min) = Rotary speed
wear_flat_length (in) = Horizontal length of cutter after wearing
d (in) = Depth of cut
rock_intrinsic_energy (lbf/in²)
zeta (unitless) = Vertical to horizontal force ratio
WOB (lb) = Weight on bit
WOB_c (lb) = Weight on bit due to cutting action
WOB_f (lb) = Weight on bit due to friction 
TOB (lb.ft) = Torque on bit
TOB_c (lb.ft) = Torque on bit due to cutting action
TOB_f (lb.ft) = Torque on bit due to friction
mu (unitless) = Friction cofficient 
d_tob (in) = Depth of cut calculated from TOB
d_wob (in) = Depth of cut calculated from WOB
dc (in) = Cutter diameter
alpha (degrees) = Back rake angle
beta (degrees) = Side rake angle
Sc (in²) = Cutting surface area
Sw (in²) = Worn surface of cutter
Rc (lbf/in²) = Index of rock strength for cutting (rock resistance)
Rp (lbf/in²) = Index of rock strength for pressing (contact pressure)
nc (unitless) = Number of cutters
sigma_t (lbf/in²) = Rock tensile strength
sigma_c (lbf/in²) = Rock uniaxial compressive strength
w (inches) =  Cutter width (same as dc in case of disc cutters)
theta_f (degrees) = Friction angle
sigma_0 (lbf/in²) = Hydrostatic stree in crushed zone
A (in) = Cut cross section area
A_ch (in) = Chamfer surface area
'''

import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os

class PDC_BIT_MODELS:
    def __init__(self, params):
        self.params = params
        # Precompute common conversions for efficiency
        self.omega = self._rpm_to_rad_per_sec(params["rotational_speed (rpm)"])
        self.bit_radius_in = params["bit_radius (inches)"]
        self.wear_flat_length_in = params["length_of_wear_flat (inches)"]

    # Conversion Methods
    def _rpm_to_rad_per_sec(self, rpm):
        return (rpm * 2 * math.pi) / 60                                                                              # Convert RPM to radians per second

    def _degrees_to_radians(self, degrees):
        return math.radians(degrees)                                                                                 # Convert degrees to radians

    def _inches_to_feet(self, inches):
        return inches / 12                                                                                           # Convert inches to feet

    # 1. Detournay's Model for WOB and TOB Calculation with bit_constant (converted to single cutter model)
    def detournay_wob_tob(self):
        params = self.params
        d = params["d (inches)"]                                                                                      # Rotational speed in rev/hr 
        WOB_c = params["zeta (unitless)"] * params["rock_intrinsic_energy (lbf/in²)"] * self.bit_radius_in * d        # lbf
        WOB_f = self.bit_radius_in * self.wear_flat_length_in * params["max_contact_pressure (lbf/in²)"]              # lbf
        WOB = WOB_c + WOB_f                                                                                           # lbf
        TOB = (params["bit_constant (unitless)"] * params["mu (unitless)"] * self.wear_flat_length_in \
               * params["max_contact_pressure (lbf/in²)"] * self.bit_radius_in**2) / 12 \
               + (params["rock_intrinsic_energy (lbf/in²)"] * self.bit_radius_in**2 * d/12)                            # lbf·ft
        return WOB, TOB

    # 2. Detournay's Model for ROP Calculation
    def detournay_rop(self):
        params = self.params
        WOB_f = self.bit_radius_in * self.wear_flat_length_in * params["max_contact_pressure (lbf/in²)"]               # lbf
        d_wob = ((params["wob (lbf)"] - WOB_f)\
                  / (params["zeta (unitless)"] * params["rock_intrinsic_energy (lbf/in²)"] * self.bit_radius_in))      # in
        TOB_f = ((params["bit_constant (unitless)"] * params["mu (unitless)"] * self.wear_flat_length_in \
                 * params["max_contact_pressure (lbf/in²)"] * self.bit_radius_in**2) / 12)                             # lbf·ft
        d_tob = ((2 * (params["tob (lbf·ft)"] - TOB_f)) \
                 / (params["rock_intrinsic_energy (lbf/in²)"] * self.bit_radius_in**2))*12                             # in
        d_avg = ((d_wob + d_tob) / 2) / 12                                                                             # Convert depth to feet
        rop = (params["rotational_speed (rpm)"] * d_wob/12 *60)                                                        # ROP in ft/hr 
        return rop, d_wob, d_tob

    # 3. Ziaja's Model for WOB, TOB, and ROP Calculation
    def ziaja_rop(self):
        params = self.params
        alpha_rad = self._degrees_to_radians(params["alpha (degrees)"])
        beta_rad = self._degrees_to_radians(params["beta (degrees)"])
        dc = params["dc (inches)"]  
        d = params["d (inches)"] 
        x = params["length_of_wear_flat (inches)"] * math.cos(alpha_rad) * math.sin(alpha_rad)                         # in (length of wear flat)
        Sc = math.cos(beta_rad) \
             * (0.24 * math.pi * dc**2 - (dc / (6 * math.cos(alpha_rad))) \
             * (4 * math.sqrt(2 * dc * math.cos(alpha_rad) * (dc * math.cos(alpha_rad) - d)) \
             - math.sqrt(16 * dc**2 * math.cos(alpha_rad)**2 - d**2)) \
             - (d / (2 * math.cos(alpha_rad)**2)) * math.sqrt(16 * dc**2 * math.cos(alpha_rad)**2 - d**2))             # in²
        Sw = (4 / 3) * x * (math.tan(alpha_rad) + 1 / math.tan(alpha_rad)) \
             * math.sqrt(dc * x * math.cos(alpha_rad) - x**2) / math.cos(alpha_rad)                                    # in²
        WOB = params["nc (unitless)"] * (params["Rc (lbf/in²)"] * Sc * math.sin(alpha_rad) \
              + params["nc (unitless)"] * params["Rp (lbf/in²)"] * Sw)                                                 # lbf
        TOB = params["r (inches)"] * params["nc (unitless)"] * (params["Rc (lbf/in²)"] * Sc * math.cos(alpha_rad) \
              + params["Rp (lbf/in²)"] * Sw * params["mu (unitless)"]) / 12                                            # lbf·ft
        ROP = params["rotational_speed (rpm)"] * self._inches_to_feet(d)* params["nc (unitless)"] * 60                 # ROP in ft/hr
        return WOB, TOB, ROP

    # 4. Che's Model for Forces and ROP Calculation
    def che_model(self):
        params = self.params
        WOB_f = self.bit_radius_in * params["max_contact_pressure (lbf/in²)"] * self.wear_flat_length_in               # lbf
        WOB_c = (params["wob (lbf)"] - WOB_f) / params["nc (unitless)"]                                                # lbf   
        alpha_rad = self._degrees_to_radians(params["alpha (degrees)"])                                                
        cutter_width = params["w (inches)"]                                                                            
        term1, term2, term3, term4 = self._che_terms(params["sigma_c (lbf/in²)"], params["sigma_t (lbf/in²)"], alpha_rad)
        d = WOB_c / (cutter_width * params["sigma_c (lbf/in²)"] * np.cos(alpha_rad) *                                  # in
                     (term1 + term2 / (4 * np.sin(np.pi / 4 - alpha_rad / 2)**2 + 
                                       4 * (params["sigma_t (lbf/in²)"] / params["sigma_c (lbf/in²)"])) + term3 + term4))
        
        ROP = ((params["rotational_speed (rpm)"] * d/12 *60)) * params["nc (unitless)"]                                # ROP in ft/hr
        return WOB_c, d, ROP

    def _che_terms(self, sigma_c, sigma_t, alpha_rad):
        term1 = -np.cos(alpha_rad) / (4 * np.sin(np.pi / 4 - alpha_rad / 2)**2 + 4 * (sigma_t / sigma_c))
        term2 = np.sqrt(np.cos(alpha_rad)**2 + 4 * (sigma_t / sigma_c) * np.sin(np.pi / 4 - alpha_rad / 2)**2 \
                + 4 * np.sin(np.pi / 4 - alpha_rad / 2)**4)
        term3 = np.sqrt(sigma_t / sigma_c + 1) / (2 * np.sin(np.pi / 4 - alpha_rad / 2))
        term4 = 0.5 * np.cos(np.pi / 4 - alpha_rad / 2) / np.sin(np.pi / 4 - alpha_rad / 2)
        return term1, term2, term3, term4

   # 5. Gerbaud's Model for Normal Force and ROP Calculation
    def gerbaud_normal_force_rop(self):
        params = self.params

        # Convert angles to radians
        alpha_rad = self._degrees_to_radians(params["alpha (degrees)"])
        theta_f_rad = self._degrees_to_radians(params["theta_f (degrees)"])

        # Calculate normal force component used in cutting action
        F_n_c = math.tan(theta_f_rad + alpha_rad) * params["A (inches²)"] * params["rock_intrinsic_energy (lbf/in²)"]  # lbf
    
        # Calculate force component on cutter chamfer used in cutting action
        F_n_ch = params["A_ch (inches²)"] *math.sin(alpha_rad)* params["sigma_0 (lbf/in²)"]                            # lbf

        # Calculate A_f similarly to Sw in Che's model for wear flat area based on Sw in Ziaja's model
        x = params["length_of_wear_flat (inches)"] * math.cos(alpha_rad) * math.sin(alpha_rad)                         # in (length of wear flat)
        dc = params["dc (inches)"]                                                                                     # in
        A_f = (4 / 3) * x * (math.tan(alpha_rad) + 1 / math.tan(alpha_rad)) \
               * math.sqrt(dc * x * math.cos(alpha_rad) - x**2) / math.cos(alpha_rad)                                  # in²

        # Calculate friction force component due to wear
        F_n_f = A_f * params["sigma_0 (lbf/in²)"]                                                                      # lbf

        # Total force used in cutting action
        F_n = F_n_c + F_n_ch + F_n_f                                                                                   # lbf

        # Calculate depth of cut (d) in inches by rearranging the wobc equation
        d = F_n* params["nc (unitless)"] \
            / (params["zeta (unitless)"] * params["rock_intrinsic_energy (lbf/in²)"] * self.bit_radius_in)             # in

        # Calculate ROP in ft/hr, converting d from inches to feet
        ROP = (d / 12) * params["rotational_speed (rpm)"] * 60 * params["nc (unitless)"]                               # ROP in ft/hr
        return F_n, d, ROP

# Function to process the Excel file, calculate model outputs, and save to an output file
def process_excel(input_file, output_file):
    # Read input file
    data = pd.read_excel(input_file)
    data.columns = [str(col).strip() for col in data.columns]
    expected_params = [
        "rotational_speed (rpm)", "bit_radius (inches)", "length_of_wear_flat (inches)", "mu (unitless)", 
        "zeta (unitless)", "rock_intrinsic_energy (lbf/in²)", "max_contact_pressure (lbf/in²)", "rop (ft/hr)", "wob (lbf)", 
        "tob (lbf·ft)", "dc (inches)", "alpha (degrees)", "beta (degrees)", "r (inches)", "nc (unitless)", "Rc (lbf/in²)", 
        "Rp (lbf/in²)", "d (inches)", "sigma_c (lbf/in²)", "sigma_t (lbf/in²)", "w (inches)", 
        "k (unitless)", "A (inches²)", "sigma_0 (lbf/in²)", "theta_f (degrees)", "A_ch (inches²)", "bit_constant (unitless)"
    ]
    
    model_row = ["Detournay"] * 5 + ["Ziaja"] * 3 + ["Che"] * 3 + ["Gerbaud"] * 3
    param_row = [
        "Detournay_WOB", "Detournay_TOB", "Detournay_ROP", "Detournay_Depth_from_WOB", "Detournay_Depth_from_TOB",
        "Ziaja_WOB", "Ziaja_TOB", "Ziaja_ROP", "Che_WOB_Cutting", "Che_Depth_of_Cut", "Che_ROP",
        "Gerbaud_Normal_Force", "Gerbaud_Depth_of_Cut", "Gerbaud_ROP"
    ]
    unit_row = ["lbf", "lbf·ft", "ft/hr", "inches", "inches",
                "lbf", "lbf·ft", "ft/hr",
                "lbf", "inches", "ft/hr",
                "lbf", "inches", "ft/hr"]

    # Prepare output data and plots
    output_data = []
    rop_input, rop_detournay, rop_ziaja, rop_che, rop_gerbaud = [], [], [], [], []
    wob_input, tob_input = [], []
    wob_detournay, tob_detournay = [], []
    wob_ziaja, tob_ziaja = [], []

    for _, row in data.iterrows():
        row_data = {param: row.get(param) for param in expected_params}
        model = PDC_BIT_MODELS(row_data)

        # Run calculations
        detournay_wob, detournay_tob = model.detournay_wob_tob()
        detournay_rop, detournay_d_wob, detournay_d_tob = model.detournay_rop()
        ziaja_wob, ziaja_tob, ziaja_rop = model.ziaja_rop()
        che_wob_c, che_d, che_rop = model.che_model()
        gerbaud_f_n, gerbaud_d, gerbaud_rop = model.gerbaud_normal_force_rop()

        # Append results to data
        output_data.append([
            detournay_wob, detournay_tob, detournay_rop, detournay_d_wob, detournay_d_tob,
            ziaja_wob, ziaja_tob, ziaja_rop,
            che_wob_c, che_d, che_rop,
            gerbaud_f_n, gerbaud_d, gerbaud_rop
        ])

        # Append for plotting
        rop_input.append(row["rop (ft/hr)"])
        rop_detournay.append(detournay_rop)
        rop_ziaja.append(ziaja_rop)
        rop_che.append(che_rop)
        rop_gerbaud.append(gerbaud_rop)

        wob_input.append(row["wob (lbf)"])
        tob_input.append(row["tob (lbf·ft)"])
        wob_detournay.append(detournay_wob)
        tob_detournay.append(detournay_tob)
        wob_ziaja.append(ziaja_wob)
        tob_ziaja.append(ziaja_tob)

    # File paths for plots
    rop_tob_plot_path = "ROP vs TOB Comparison.png"
    rop_wob_plot_path = "ROP vs WOB Comparison.png"
    wob_tob_plot_path = "WOB vs TOB Comparison.png"

    
    # Plot 1: ROP vs TOB Comparison
    plt.figure(figsize=(10, 6))
    plt.plot(tob_input,rop_detournay, label="ROP (Detournay)", marker="x", linestyle="")
    plt.plot(tob_input, rop_ziaja, label="ROP (Ziaja)", marker="s", linestyle="")
    plt.plot(tob_input, rop_che, label="ROP (Che)", marker="d", linestyle="")
    plt.plot(tob_input, rop_gerbaud, label="ROP (Gerbaud)", marker="^", linestyle="")
    plt.plot(tob_input, rop_input, label="ROP (Real)", marker="+", linestyle="")
    detournay_fit = np.polyfit(tob_input, rop_detournay, 1)  # Linear fit (degree 1)
    ziaja_fit = np.polyfit(tob_input, rop_ziaja, 1)
    che_fit = np.polyfit(tob_input, rop_che, 1)
    gerbaud_fit = np.polyfit(tob_input, rop_gerbaud, 1)
    real_fit = np.polyfit(tob_input, rop_input, 1)
    plt.plot(tob_input, np.polyval(detournay_fit, tob_input), label="Trend Detournay", linestyle="--")
    plt.plot(tob_input, np.polyval(ziaja_fit, tob_input), label="Trend Ziaja", linestyle="--")
    plt.plot(tob_input, np.polyval(che_fit, tob_input), label="Trend Che", linestyle="--")
    plt.plot(tob_input, np.polyval(gerbaud_fit, tob_input), label="Trend Gerbaud", linestyle="--")
    plt.plot(tob_input, np.polyval(real_fit, tob_input), label="Trend Real", linestyle="--")
    plt.legend()
    plt.xlabel("TOB (Ibf.ft)")
    plt.ylabel("ROP (ft/hr)")
    plt.title("ROP vs TOB")
    plt.grid()
    plt.savefig(rop_tob_plot_path)
    plt.show()

    # Plot 2: ROP vs WOB Comparison
    plt.figure(figsize=(10, 6))
    plt.plot(wob_input,rop_detournay, label="ROP (Detournay)", marker="x", linestyle="")
    plt.plot(wob_input, rop_ziaja, label="ROP (Ziaja)", marker="s", linestyle="")
    plt.plot(wob_input, rop_che, label="ROP (Che)", marker="d", linestyle="")
    plt.plot(wob_input, rop_gerbaud, label="ROP (Gerbaud)", marker="^", linestyle="")
    plt.plot(wob_input, rop_input, label="ROP (Real)", marker="+", linestyle="")
    detournay_fit = np.polyfit(wob_input, rop_detournay, 1)  # Linear fit (degree 1)
    ziaja_fit = np.polyfit(wob_input, rop_ziaja, 1)
    che_fit = np.polyfit(wob_input, rop_che, 1)
    gerbaud_fit = np.polyfit(wob_input, rop_gerbaud, 1)
    real_fit = np.polyfit(wob_input, rop_input, 1)
    plt.plot(wob_input, np.polyval(detournay_fit, wob_input), label="Trend Detournay", linestyle="--")
    plt.plot(wob_input, np.polyval(ziaja_fit, wob_input), label="Trend Ziaja", linestyle="--")
    plt.plot(wob_input, np.polyval(che_fit, wob_input), label="Trend Che", linestyle="--")
    plt.plot(wob_input, np.polyval(gerbaud_fit, wob_input), label="Trend Gerbaud", linestyle="--")
    plt.plot(wob_input, np.polyval(real_fit, wob_input), label="Trend Real", linestyle="--")
    plt.legend()
    plt.xlabel("WOB (Ibf)")
    plt.ylabel("ROP (ft/hr)")
    plt.title("ROP vs WOB")
    plt.grid()
    plt.savefig(rop_wob_plot_path)
    plt.show()

    # Plot 3: WOB vs TOB Comparison
    plt.figure(figsize=(10, 6))
    plt.plot(tob_detournay, wob_detournay, label="Detournay", marker="x", linestyle="")
    plt.plot(tob_ziaja, wob_ziaja, label="Ziaja", marker="s", linestyle="")
    plt.plot(tob_input, wob_input, label="Real", marker="+", linestyle="")
    detournay_fit = np.polyfit(tob_detournay, wob_detournay, 1)  # Linear fit (degree 1)
    ziaja_fit = np.polyfit(tob_ziaja, wob_ziaja, 1)
    real_fit = np.polyfit(tob_input, wob_input, 1)
    plt.plot(tob_detournay, np.polyval(detournay_fit, tob_detournay), label="Trend Detournay", linestyle="--")
    plt.plot(tob_ziaja, np.polyval(ziaja_fit, tob_ziaja), label="Trend Ziaja", linestyle="--")
    plt.plot(tob_input, np.polyval(real_fit, tob_input), label="Trend Real", linestyle="--")
    plt.legend()
    plt.xlabel("TOB (Ibf.ft)")
    plt.ylabel("WOB (Ibf)")
    plt.title("WOB vs TOB Comparison")
    plt.grid()
    plt.savefig(wob_tob_plot_path)
    plt.show()

    from openpyxl import load_workbook
    # Save results and plots to the Excel file
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        # Remove existing sheets
        if "Results" in wb.sheetnames:
            del wb["Results"]
        if "Plots" in wb.sheetnames:
            del wb["Plots"]
    else:
        wb = Workbook()

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    
    # Add a new "Plots" sheet
    ws = wb.create_sheet(title="Plots")
    ws.add_image(Image(rop_tob_plot_path), "B2")
    ws.add_image(Image(rop_wob_plot_path), "R2")
    ws.add_image(Image(wob_tob_plot_path), "AH2")
    wb.save(output_file)

    # Write "Results" sheet using pandas
    results_df = pd.DataFrame(output_data, columns=param_row)
    results_df.loc[-1] = model_row  # Insert model names at the first row
    results_df.loc[-2] = param_row  # Insert parameter names at the second row
    results_df.loc[-3] = unit_row  # Insert units at the third row
    results_df.index = results_df.index + 3  # Shift index for new rows
    results_df.sort_index(inplace=True)

    with pd.ExcelWriter(output_file, mode="a", engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="Results", index=False)

        
